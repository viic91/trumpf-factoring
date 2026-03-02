"""
Excel-Export-Modul fuer das Trumpf Factoring Tool.

Erzeugt professionelle Excel-Reports mit Dashboard, offenen Posten,
Gesamtuebersicht und Zinsanalyse.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Farb-Konstanten
# ---------------------------------------------------------------------------
COLOR_HEADER = "1B2A4A"
COLOR_ACCENT = "2E86AB"
COLOR_SUCCESS = "28A745"
COLOR_WARNING = "FFC107"
COLOR_DANGER = "DC3545"
COLOR_ALT_ROW = "F8F9FA"
COLOR_WHITE = "FFFFFF"

FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
FILL_ACCENT = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type="solid")
FILL_SUCCESS = PatternFill(start_color=COLOR_SUCCESS, end_color=COLOR_SUCCESS, fill_type="solid")
FILL_WARNING = PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type="solid")
FILL_DANGER = PatternFill(start_color=COLOR_DANGER, end_color=COLOR_DANGER, fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type="solid")
FILL_WHITE = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

FONT_HEADER = Font(bold=True, color=COLOR_WHITE, size=11)
FONT_TITLE = Font(bold=True, color=COLOR_HEADER, size=18)
FONT_KPI_LABEL = Font(bold=True, color=COLOR_HEADER, size=11)
FONT_KPI_VALUE = Font(bold=True, color=COLOR_ACCENT, size=14)
FONT_SUBTITLE = Font(bold=True, color=COLOR_HEADER, size=12)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CURRENCY_FORMAT = '#,##0.00 "EUR"'
PERCENT_FORMAT = "0.00%"
NUMBER_FORMAT = "#,##0.00"
INTEGER_FORMAT = "#,##0"


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def _apply_header_style(ws, row: int, max_col: int) -> None:
    """Formatiert eine Kopfzeile mit dunklem Hintergrund und weisser Schrift."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 30) -> None:
    """Passt die Spaltenbreiten automatisch an den Inhalt an."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    """Schreibt einen DataFrame in ein Worksheet und gibt die naechste freie Zeile zurueck."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
    return start_row + len(df) + 1  # +1 fuer Header


def _apply_alternating_rows(ws, start_row: int, end_row: int, max_col: int) -> None:
    """Wendet abwechselnde Zeilenfarben an (weiss / hellgrau)."""
    for row_idx in range(start_row, end_row + 1):
        fill = FILL_ALT_ROW if (row_idx - start_row) % 2 == 1 else FILL_WHITE
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill


def _apply_currency_format(ws, col: int, start_row: int, end_row: int) -> None:
    """Wendet das Waehrungsformat auf eine Spalte an."""
    for row_idx in range(start_row, end_row + 1):
        ws.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT


def _apply_percent_format(ws, col: int, start_row: int, end_row: int) -> None:
    """Wendet das Prozentformat auf eine Spalte an."""
    for row_idx in range(start_row, end_row + 1):
        ws.cell(row=row_idx, column=col).number_format = PERCENT_FORMAT


def _apply_number_format(ws, col: int, start_row: int, end_row: int, fmt: str = NUMBER_FORMAT) -> None:
    """Wendet ein Zahlenformat auf eine Spalte an."""
    for row_idx in range(start_row, end_row + 1):
        ws.cell(row=row_idx, column=col).number_format = fmt


# ---------------------------------------------------------------------------
# Sheet-Builder
# ---------------------------------------------------------------------------

def _build_dashboard(wb: Workbook, df: pd.DataFrame) -> None:
    """Erstellt das Dashboard-Sheet mit KPIs und Diagrammen."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = COLOR_HEADER

    # Titel
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "TRUMPF Factoring Report"
    title_cell.font = FONT_TITLE
    title_cell.alignment = Alignment(vertical="center")

    # Datum
    ws.merge_cells("A2:H2")
    date_cell = ws["A2"]
    date_cell.value = f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M Uhr')}"
    date_cell.font = Font(italic=True, color="666666", size=10)

    # --- KPIs berechnen ---
    offen_df = df[df["status"] != "Abgeschlossen"]
    offene_summe = offen_df["offener_betrag"].sum() if not offen_df.empty else 0
    avg_jahreszins = df["eff_jahreszins"].mean() if not df.empty else 0
    avg_tage = df["tage_finanziert"].mean() if not df.empty else 0
    anzahl_offen = len(offen_df)

    kpi_data = [
        ("Offene Betraege Gesamt", f"{offene_summe:,.2f} EUR"),
        ("Durchschn. eff. Jahreszins", f"{avg_jahreszins:.2%}"),
        ("Durchschn. Tage finanziert", f"{avg_tage:.1f}"),
        ("Anzahl offene Rechnungen", f"{anzahl_offen}"),
    ]

    # KPI-Boxen in Zeile 4-6
    for i, (label, value) in enumerate(kpi_data):
        col_start = 1 + i * 2
        col_end = col_start + 1
        col_start_letter = get_column_letter(col_start)
        col_end_letter = get_column_letter(col_end)

        # Label
        ws.merge_cells(f"{col_start_letter}4:{col_end_letter}4")
        label_cell = ws.cell(row=4, column=col_start, value=label)
        label_cell.font = FONT_KPI_LABEL
        label_cell.alignment = Alignment(horizontal="center")

        # Wert
        ws.merge_cells(f"{col_start_letter}5:{col_end_letter}5")
        value_cell = ws.cell(row=5, column=col_start, value=value)
        value_cell.font = FONT_KPI_VALUE
        value_cell.alignment = Alignment(horizontal="center")

        # Rahmen fuer die Box
        for r in (4, 5, 6):
            for c in (col_start, col_end):
                ws.cell(row=r, column=c).border = THIN_BORDER
        # Trennlinie unten
        ws.merge_cells(f"{col_start_letter}6:{col_end_letter}6")
        accent_cell = ws.cell(row=6, column=col_start)
        accent_cell.fill = FILL_ACCENT

    # --- Balkendiagramm: Offene Betraege nach Valuta-Monat ---
    if not offen_df.empty and "valuta_trumpf" in offen_df.columns:
        chart_df = offen_df.copy()
        chart_df["valuta_trumpf"] = pd.to_datetime(chart_df["valuta_trumpf"], errors="coerce")
        chart_df = chart_df.dropna(subset=["valuta_trumpf"])

        if not chart_df.empty:
            chart_df["monat"] = chart_df["valuta_trumpf"].dt.to_period("M").astype(str)
            monthly = chart_df.groupby("monat")["offener_betrag"].sum().reset_index()
            monthly.columns = ["Monat", "Offener Betrag"]
            monthly = monthly.sort_values("Monat")

            # Daten in ein Hilfsbereich schreiben (ab Zeile 9)
            ws.cell(row=8, column=1, value="Offene Betraege nach Valuta-Monat").font = FONT_SUBTITLE
            ws.cell(row=9, column=1, value="Monat").font = FONT_HEADER
            ws.cell(row=9, column=1).fill = FILL_HEADER
            ws.cell(row=9, column=2, value="Offener Betrag").font = FONT_HEADER
            ws.cell(row=9, column=2).fill = FILL_HEADER

            for idx, (_, row_data) in enumerate(monthly.iterrows()):
                ws.cell(row=10 + idx, column=1, value=row_data["Monat"])
                ws.cell(row=10 + idx, column=2, value=row_data["Offener Betrag"])
                ws.cell(row=10 + idx, column=2).number_format = CURRENCY_FORMAT

            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.title = "Offene Betraege nach Valuta-Monat"
            bar_chart.y_axis.title = "Betrag (EUR)"
            bar_chart.x_axis.title = "Monat"
            bar_chart.style = 10

            data_ref = Reference(ws, min_col=2, min_row=9, max_row=9 + len(monthly))
            cats_ref = Reference(ws, min_col=1, min_row=10, max_row=9 + len(monthly))
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            bar_chart.shape = 4
            bar_chart.width = 18
            bar_chart.height = 12

            chart_start_row = 10 + len(monthly) + 2
            ws.add_chart(bar_chart, f"A{chart_start_row}")
            pie_anchor_row = chart_start_row
        else:
            pie_anchor_row = 10
    else:
        pie_anchor_row = 10

    # --- Kreisdiagramm: Verteilung nach Lieferant ---
    if not offen_df.empty:
        supplier_data = offen_df.groupby("lieferant")["offener_betrag"].sum().reset_index()
        supplier_data.columns = ["Lieferant", "Offener Betrag"]
        supplier_data = supplier_data.sort_values("Offener Betrag", ascending=False)

        pie_data_start = pie_anchor_row + 18  # Platz fuer Balkendiagramm
        ws.cell(row=pie_data_start - 1, column=1, value="Verteilung nach Lieferant").font = FONT_SUBTITLE
        ws.cell(row=pie_data_start, column=1, value="Lieferant").font = FONT_HEADER
        ws.cell(row=pie_data_start, column=1).fill = FILL_HEADER
        ws.cell(row=pie_data_start, column=2, value="Offener Betrag").font = FONT_HEADER
        ws.cell(row=pie_data_start, column=2).fill = FILL_HEADER

        for idx, (_, row_data) in enumerate(supplier_data.iterrows()):
            ws.cell(row=pie_data_start + 1 + idx, column=1, value=row_data["Lieferant"])
            ws.cell(row=pie_data_start + 1 + idx, column=2, value=row_data["Offener Betrag"])
            ws.cell(row=pie_data_start + 1 + idx, column=2).number_format = CURRENCY_FORMAT

        pie_chart = PieChart()
        pie_chart.title = "Verteilung nach Lieferant"
        pie_chart.style = 10
        pie_chart.width = 16
        pie_chart.height = 12

        data_ref = Reference(ws, min_col=2, min_row=pie_data_start, max_row=pie_data_start + len(supplier_data))
        cats_ref = Reference(ws, min_col=1, min_row=pie_data_start + 1, max_row=pie_data_start + len(supplier_data))
        pie_chart.add_data(data_ref, titles_from_data=True)
        pie_chart.set_categories(cats_ref)

        chart_row = pie_data_start + len(supplier_data) + 2
        ws.add_chart(pie_chart, f"D{pie_anchor_row + 18}")

    _auto_width(ws)


def _build_offene_posten(wb: Workbook, df: pd.DataFrame) -> None:
    """Erstellt das Sheet 'Offene Posten' mit bedingter Formatierung."""
    ws = wb.create_sheet("Offene Posten")
    ws.sheet_properties.tabColor = COLOR_DANGER

    offen_df = df[df["status"] != "Abgeschlossen"].copy()

    # Spaltenauswahl und deutsche Bezeichnungen
    column_map = {
        "lieferant": "Lieferant",
        "re_nr_lieferant": "RE-Nr.",
        "valuta_trumpf": "Valuta",
        "trumpf_brutto": "Trumpf Brutto",
        "bereits_gezahlt": "Bereits gezahlt",
        "offener_betrag": "Offen",
        "status": "Status",
    }

    display_df = offen_df[list(column_map.keys())].copy()
    display_df.columns = list(column_map.values())

    # Daten schreiben
    _write_dataframe(ws, display_df, start_row=1)
    header_row = 1
    data_start = 2
    data_end = data_start + len(display_df) - 1
    max_col = len(column_map)

    # Header-Stil
    _apply_header_style(ws, header_row, max_col)

    # Waehrungsformat fuer Geldspalten (Spalte 4, 5, 6)
    currency_cols = [4, 5, 6]
    for col in currency_cols:
        _apply_currency_format(ws, col, data_start, max(data_end, data_start))

    # Bedingte Formatierung auf Spalte "Offen" (Spalte 6)
    offen_col_letter = get_column_letter(6)
    cell_range = f"{offen_col_letter}{data_start}:{offen_col_letter}{data_end + 100}"

    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["20000"],
            fill=PatternFill(start_color=COLOR_DANGER, end_color=COLOR_DANGER, fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between",
            formula=["10000.01", "20000"],
            fill=PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["10000"],
            fill=PatternFill(start_color=COLOR_SUCCESS, end_color=COLOR_SUCCESS, fill_type="solid"),
        ),
    )

    # Autofilter
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max(data_end, data_start)}"

    _auto_width(ws)


def _build_alle_rechnungen(wb: Workbook, df: pd.DataFrame) -> None:
    """Erstellt das Sheet 'Alle Rechnungen' mit allen relevanten Spalten."""
    ws = wb.create_sheet("Alle Rechnungen")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    # Relevante Spalten auswaehlen (id, Pfade und Timestamps ueberspringen)
    skip_cols = {"id"}
    column_map = {
        "lieferant": "Lieferant",
        "re_nr_lieferant": "RE-Nr. Lieferant",
        "re_datum_lieferant": "RE-Datum Lieferant",
        "netto_betrag": "Netto Betrag",
        "brutto_betrag": "Brutto Betrag",
        "valuta_trumpf": "Valuta Trumpf",
        "re_nr_trumpf": "RE-Nr. Trumpf",
        "re_datum_trumpf": "RE-Datum Trumpf",
        "zahlung_an_trumpf": "Zahlung an Trumpf",
        "trumpf_netto": "Trumpf Netto",
        "trumpf_brutto": "Trumpf Brutto",
        "bereits_gezahlt": "Bereits gezahlt",
        "offener_betrag": "Offener Betrag",
        "status": "Status",
        "zinsen": "Zinsen",
        "zinsaufschlag": "Zinsaufschlag",
        "eff_jahreszins": "Eff. Jahreszins",
        "tage_finanziert": "Tage finanziert",
        "zinsen_pro_tag": "Zinsen pro Tag",
        "zinssatz_30_tage": "Zinssatz 30 Tage",
    }

    available_cols = [c for c in column_map if c in df.columns and c not in skip_cols]
    display_df = df[available_cols].copy()
    display_df.columns = [column_map[c] for c in available_cols]

    # Daten schreiben
    _write_dataframe(ws, display_df, start_row=1)
    header_row = 1
    data_start = 2
    data_end = data_start + len(display_df) - 1
    max_col = len(available_cols)

    # Header-Stil
    _apply_header_style(ws, header_row, max_col)

    # Alternating rows
    _apply_alternating_rows(ws, data_start, max(data_end, data_start), max_col)

    # Waehrungsformate
    currency_columns = [
        "Netto Betrag", "Brutto Betrag", "Trumpf Netto", "Trumpf Brutto",
        "Bereits gezahlt", "Offener Betrag", "Zinsen", "Zinsaufschlag",
        "Zahlung an Trumpf", "Zinsen pro Tag",
    ]
    german_headers = [column_map[c] for c in available_cols]
    for col_name in currency_columns:
        if col_name in german_headers:
            col_idx = german_headers.index(col_name) + 1
            _apply_currency_format(ws, col_idx, data_start, max(data_end, data_start))

    # Prozentformat fuer Jahreszins
    percent_columns = ["Eff. Jahreszins", "Zinssatz 30 Tage"]
    for col_name in percent_columns:
        if col_name in german_headers:
            col_idx = german_headers.index(col_name) + 1
            _apply_percent_format(ws, col_idx, data_start, max(data_end, data_start))

    # Ganzzahlformat fuer Tage
    if "Tage finanziert" in german_headers:
        col_idx = german_headers.index("Tage finanziert") + 1
        _apply_number_format(ws, col_idx, data_start, max(data_end, data_start), fmt=INTEGER_FORMAT)

    # Autofilter
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max(data_end, data_start)}"

    # Erste Zeile fixieren
    ws.freeze_panes = "A2"

    _auto_width(ws)


def _build_zinsanalyse(wb: Workbook, df: pd.DataFrame) -> None:
    """Erstellt das Sheet 'Zinsanalyse' mit Zusammenfassung nach Lieferant."""
    ws = wb.create_sheet("Zinsanalyse")
    ws.sheet_properties.tabColor = COLOR_WARNING

    # Gruppierung nach Lieferant
    if df.empty:
        summary = pd.DataFrame(columns=[
            "Lieferant", "Anzahl Rechnungen", "Gesamtvolumen Netto",
            "Gesamtzinsen", "Durchschn. Jahreszins", "Durchschn. Tage",
        ])
    else:
        grouped = df.groupby("lieferant").agg(
            anzahl=("lieferant", "count"),
            volumen_netto=("netto_betrag", "sum"),
            gesamt_zinsen=("zinsen", "sum"),
            avg_jahreszins=("eff_jahreszins", "mean"),
            avg_tage=("tage_finanziert", "mean"),
        ).reset_index()

        summary = pd.DataFrame({
            "Lieferant": grouped["lieferant"],
            "Anzahl Rechnungen": grouped["anzahl"],
            "Gesamtvolumen Netto": grouped["volumen_netto"],
            "Gesamtzinsen": grouped["gesamt_zinsen"],
            "Durchschn. Jahreszins": grouped["avg_jahreszins"],
            "Durchschn. Tage": grouped["avg_tage"],
        })

    # Daten schreiben
    _write_dataframe(ws, summary, start_row=1)
    header_row = 1
    data_start = 2
    data_end = data_start + len(summary) - 1
    max_col = 6

    # Header-Stil
    _apply_header_style(ws, header_row, max_col)

    # Alternating rows
    if not summary.empty:
        _apply_alternating_rows(ws, data_start, data_end, max_col)

    # Formate
    _apply_number_format(ws, 2, data_start, max(data_end, data_start), fmt=INTEGER_FORMAT)
    _apply_currency_format(ws, 3, data_start, max(data_end, data_start))
    _apply_currency_format(ws, 4, data_start, max(data_end, data_start))
    _apply_percent_format(ws, 5, data_start, max(data_end, data_start))
    _apply_number_format(ws, 6, data_start, max(data_end, data_start), fmt="0.0")

    # Totalzeile
    if not summary.empty:
        total_row = data_end + 1
        ws.cell(row=total_row, column=1, value="GESAMT").font = Font(bold=True, color=COLOR_HEADER, size=11)
        ws.cell(row=total_row, column=2, value=summary["Anzahl Rechnungen"].sum())
        ws.cell(row=total_row, column=2).number_format = INTEGER_FORMAT
        ws.cell(row=total_row, column=3, value=summary["Gesamtvolumen Netto"].sum())
        ws.cell(row=total_row, column=3).number_format = CURRENCY_FORMAT
        ws.cell(row=total_row, column=4, value=summary["Gesamtzinsen"].sum())
        ws.cell(row=total_row, column=4).number_format = CURRENCY_FORMAT
        ws.cell(row=total_row, column=5, value=summary["Durchschn. Jahreszins"].mean())
        ws.cell(row=total_row, column=5).number_format = PERCENT_FORMAT
        ws.cell(row=total_row, column=6, value=summary["Durchschn. Tage"].mean())
        ws.cell(row=total_row, column=6).number_format = "0.0"

        # Totalzeile formatieren
        for col in range(1, max_col + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color=COLOR_HEADER, size=11)
            cell.border = Border(top=Side(style="double"), bottom=Side(style="double"))

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Hauptfunktion
# ---------------------------------------------------------------------------

def generate_excel_report(df: pd.DataFrame) -> bytes:
    """
    Erzeugt einen professionellen Excel-Report fuer das Trumpf Factoring Tool.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame mit allen Rechnungsdaten. Erwartete Spalten:
        id, lieferant, re_nr_lieferant, re_datum_lieferant, netto_betrag,
        brutto_betrag, valuta_trumpf, re_nr_trumpf, re_datum_trumpf,
        zahlung_an_trumpf, trumpf_netto, trumpf_brutto, bereits_gezahlt,
        offener_betrag, status, zinsen, zinsaufschlag, eff_jahreszins,
        tage_finanziert, zinsen_pro_tag, zinssatz_30_tage

    Returns
    -------
    bytes
        Excel-Datei als Bytes (z.B. fuer Streamlit st.download_button).
    """
    wb = Workbook()

    _build_dashboard(wb, df)
    _build_offene_posten(wb, df)
    _build_alle_rechnungen(wb, df)
    _build_zinsanalyse(wb, df)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
